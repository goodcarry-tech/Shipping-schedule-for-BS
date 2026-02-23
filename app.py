"""
船期整理系統 - Shipping Schedule Organizer
支援: CNC (上傳), TSL (上傳), IAL (網頁爬取), KMTC (網頁爬取), YML (網頁爬取)
POL: HAIPHONG | POD: HONG KONG / SHEKOU / KAOHSIUNG / TAICHUNG
"""

import streamlit as st
import pandas as pd
import io
import re
import os
import time
import warnings
from datetime import datetime, timedelta
from copy import copy

warnings.filterwarnings("ignore")

try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ─────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────
POD_LIST    = ["HONG KONG", "SHEKOU", "KAOHSIUNG", "TAICHUNG"]
OUTPUT_COLS = ["POL", "POD", "Vessel", "Voyage", "ETD", "ETA", "T/T Time", "CY Cut-off", "SI Cut-off"]
CARRIER_LIST = ["CNC", "TSL", "IAL", "KMTC", "YML"]

DAY_TO_NUM = {"MON": 0, "TUE": 1, "WED": 2, "THU": 3, "FRI": 4, "SAT": 5, "SUN": 6}

# ─────────────────────────────────────────────
# Helper utilities
# ─────────────────────────────────────────────

def normalize_pod(pod_str: str) -> str:
    """Map raw POD string to standard form."""
    s = pod_str.strip().upper()
    for std in POD_LIST:
        if std in s:
            return std
    return s

def safe_date_str(val) -> str:
    """Convert various date types to YYYY/MM/DD string."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, (datetime,)):
        return val.strftime("%Y/%m/%d")
    if isinstance(val, str):
        val = val.strip()
        # Try common formats
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y/%m/%d")
            except ValueError:
                pass
        return val
    try:
        import pandas as pd
        if pd.notna(val):
            return pd.Timestamp(val).strftime("%Y/%m/%d")
    except Exception:
        pass
    return str(val)

def days_back_from_etd(etd_date: datetime, day_name: str) -> datetime:
    """
    Given ETD date and a day name (e.g. 'SAT'), return the
    most recent occurrence of that day BEFORE or ON etd_date.
    """
    day_name = day_name.strip().upper()[:3]
    target_wd = DAY_TO_NUM.get(day_name)
    if target_wd is None:
        return etd_date
    etd_wd = etd_date.weekday()
    delta = (etd_wd - target_wd) % 7
    return etd_date - timedelta(days=delta)

def empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=OUTPUT_COLS)

# ─────────────────────────────────────────────
# CNC Parser
# ─────────────────────────────────────────────

def parse_cnc(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Parse CNC schedule file (CSV or PDF)."""
    rows = []
    fname_lower = filename.lower()

    if fname_lower.endswith(".csv"):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes))
            df.columns = [str(c).strip() for c in df.columns]

            col_map = {
                "Origin":       "POL",
                "Destination":  "POD",
                "Vessel name":  "Vessel",
                "Vessel Name":  "Vessel",
                "Voyage ref.":  "Voyage",
                "Voyage Ref.":  "Voyage",
                "Voyage ref":   "Voyage",
                "Departure Date": "ETD",
                "Arrival Date":   "ETA",
                "Transit Time":   "T/T Time",
                "Port cut-off":   "CY Cut-off",
                "Port Cut-off":   "CY Cut-off",
                "SI cut-off":     "SI Cut-off",
                "SI Cut-off":     "SI Cut-off",
            }
            df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

            for _, row in df.iterrows():
                pol = str(row.get("POL", "")).strip().upper()
                pod = normalize_pod(str(row.get("POD", "")))
                if pod not in POD_LIST:
                    continue
                if "HAIPHONG" not in pol:
                    continue

                tt = str(row.get("T/T Time", "")).strip()
                if tt and not tt.lower().endswith("day") and not tt.lower().endswith("days"):
                    tt = f"{tt} days"

                rows.append({
                    "POL":        "HAIPHONG",
                    "POD":        pod,
                    "Vessel":     str(row.get("Vessel", "")).strip(),
                    "Voyage":     str(row.get("Voyage", "")).strip(),
                    "ETD":        safe_date_str(row.get("ETD", "")),
                    "ETA":        safe_date_str(row.get("ETA", "")),
                    "T/T Time":   tt,
                    "CY Cut-off": str(row.get("CY Cut-off", "")).strip(),
                    "SI Cut-off": str(row.get("SI Cut-off", "")).strip(),
                })
        except Exception as e:
            st.error(f"CNC CSV parse error: {e}")

    elif fname_lower.endswith(".pdf"):
        if not PDF_AVAILABLE:
            st.error("pdfplumber is required to parse PDFs. Run: pip install pdfplumber")
            return empty_df()
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        # Find header row
                        header_idx = None
                        header = []
                        for i, row in enumerate(table):
                            row_text = " ".join([str(c or "") for c in row]).lower()
                            if "vessel" in row_text or "voyage" in row_text or "departure" in row_text:
                                header_idx = i
                                header = [str(c or "").strip() for c in row]
                                break
                        if header_idx is None:
                            continue

                        col_map_pdf = {}
                        for j, h in enumerate(header):
                            hl = h.lower()
                            if "origin" in hl:        col_map_pdf["POL"]        = j
                            elif "destination" in hl: col_map_pdf["POD"]        = j
                            elif "vessel name" in hl: col_map_pdf["Vessel"]     = j
                            elif "voyage" in hl:      col_map_pdf["Voyage"]     = j
                            elif "departure" in hl:   col_map_pdf["ETD"]        = j
                            elif "arrival date" in hl: col_map_pdf["ETA"]       = j
                            elif "transit" in hl:     col_map_pdf["T/T Time"]   = j
                            elif "port cut" in hl:    col_map_pdf["CY Cut-off"] = j
                            elif "si cut" in hl:      col_map_pdf["SI Cut-off"] = j

                        for row in table[header_idx + 1:]:
                            def g(key):
                                idx = col_map_pdf.get(key)
                                return str(row[idx] or "").strip() if idx is not None and idx < len(row) else ""

                            pol = g("POL").upper()
                            pod = normalize_pod(g("POD"))
                            if pod not in POD_LIST:
                                continue

                            tt = g("T/T Time")
                            if tt and not tt.lower().endswith("day") and not tt.lower().endswith("days"):
                                tt = f"{tt} days"

                            rows.append({
                                "POL":        "HAIPHONG",
                                "POD":        pod,
                                "Vessel":     g("Vessel"),
                                "Voyage":     g("Voyage"),
                                "ETD":        safe_date_str(g("ETD")),
                                "ETA":        safe_date_str(g("ETA")),
                                "T/T Time":   tt,
                                "CY Cut-off": g("CY Cut-off"),
                                "SI Cut-off": g("SI Cut-off"),
                            })
        except Exception as e:
            st.error(f"CNC PDF parse error: {e}")

    else:
        # Try Excel
        try:
            df = pd.read_excel(io.BytesIO(file_bytes))
            df.columns = [str(c).strip() for c in df.columns]
            col_map = {
                "Origin": "POL", "Destination": "POD",
                "Vessel name": "Vessel", "Vessel Name": "Vessel",
                "Voyage ref.": "Voyage", "Voyage ref": "Voyage",
                "Departure Date": "ETD", "Arrival Date": "ETA",
                "Transit Time": "T/T Time",
                "Port cut-off": "CY Cut-off",
                "SI cut-off": "SI Cut-off",
            }
            df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)
            for _, row in df.iterrows():
                pod = normalize_pod(str(row.get("POD", "")))
                if pod not in POD_LIST:
                    continue
                tt = str(row.get("T/T Time", "")).strip()
                if tt and not tt.lower().endswith("day") and not tt.lower().endswith("days"):
                    tt = f"{tt} days"
                rows.append({
                    "POL": "HAIPHONG", "POD": pod,
                    "Vessel": str(row.get("Vessel", "")).strip(),
                    "Voyage": str(row.get("Voyage", "")).strip(),
                    "ETD": safe_date_str(row.get("ETD", "")),
                    "ETA": safe_date_str(row.get("ETA", "")),
                    "T/T Time": tt,
                    "CY Cut-off": str(row.get("CY Cut-off", "")).strip(),
                    "SI Cut-off": str(row.get("SI Cut-off", "")).strip(),
                })
        except Exception as e:
            st.error(f"CNC Excel parse error: {e}")

    return pd.DataFrame(rows, columns=OUTPUT_COLS) if rows else empty_df()


# ─────────────────────────────────────────────
# TSL Parser
# ─────────────────────────────────────────────

def parse_tsl(file_bytes: bytes, filename: str,
              start_date: datetime, end_date: datetime) -> pd.DataFrame:
    """
    Parse TSL schedule file (Excel or PDF).
    Handles day-based CY cut-off and SI cut-off calculation.
    """
    rows = []
    fname_lower = filename.lower()

    def _parse_tsl_excel(wb):
        for ws in wb.worksheets:
            _parse_tsl_sheet(ws, rows, start_date, end_date)

    def _parse_tsl_sheet(ws, out_rows, start_dt, end_dt):
        # Read all cells into a 2D list for easy traversal
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(c).strip() if c is not None else "" for c in row])

        if not data:
            return

        # ── Find the main header row ──────────────────────────────────────
        header_row_idx = None
        header = []
        for i, row in enumerate(data):
            row_upper = [c.upper() for c in row]
            if ("VESSEL" in row_upper or "VESSEL NAME" in " ".join(row_upper)) and \
               ("VOYAGE" in row_upper or any("VOYAGE" in c for c in row_upper)):
                header_row_idx = i
                header = row_upper
                break

        if header_row_idx is None:
            # Try to find any row with ETD
            for i, row in enumerate(data):
                row_upper = [c.upper() for c in row]
                if any("ETD" in c for c in row_upper) and any("ETA" in c for c in row_upper):
                    header_row_idx = i
                    header = row_upper
                    break

        if header_row_idx is None:
            return

        # ── Identify column indices ───────────────────────────────────────
        def find_col(keywords):
            for j, h in enumerate(header):
                if any(kw.upper() in h for kw in keywords):
                    return j
            return None

        c_service = find_col(["SERVICE"])
        c_vessel  = find_col(["VESSEL"])
        c_voyage  = find_col(["VOYAGE"])
        c_etd     = find_col(["ETD"])
        c_eta     = find_col(["ETA"])

        # POD columns: columns with POD port names in header
        pod_cols = {}
        for j, h in enumerate(header):
            for pod in POD_LIST:
                if pod in h:
                    pod_cols[pod] = j
                    break

        if c_vessel is None or c_voyage is None or c_etd is None:
            return

        # ── Look for SERVICE reference table (cut-off rules) ─────────────
        # Format: SERVICE NAME | ETD day | CY time+day | SI time+day
        service_rules = {}  # {service_name: {"cy_day": str, "cy_time": str, "si_day": str, "si_time": str}}

        # Search for a "SERVICE NAME" section anywhere before/after data
        for i, row in enumerate(data):
            row_upper = [c.upper() for c in row]
            if "SERVICE NAME" in " ".join(row_upper) or any("SERVICE NAME" in c for c in row_upper):
                # This row or next rows contain the reference table
                svc_header_idx = i
                # Column positions in this section
                svc_col = next((j for j, c in enumerate(row_upper) if "SERVICE" in c and "NAME" in c), 0)
                etd_hph_col = next((j for j, c in enumerate(row_upper) if "ETD" in c and ("HPH" in c or "HAI" in c or "DEP" in c)), None)
                cy_col  = next((j for j, c in enumerate(row_upper) if "CY" in c and "CUT" in c.replace(" ", "") or c == "CY"), None)
                si_col  = next((j for j, c in enumerate(row_upper) if "SUBMIT" in c or ("SI" in c and "VGM" in c) or c == "SI/VGM"), None)

                if etd_hph_col is None:
                    etd_hph_col = next((j for j, c in enumerate(row_upper) if "ETD HPH" in c or "ETD" in c), None)
                if cy_col is None:
                    cy_col = next((j for j, c in enumerate(row_upper) if "CY" in c), None)
                if si_col is None:
                    si_col = next((j for j, c in enumerate(row_upper) if "SI" in c or "VGM" in c), None)

                # Read subsequent rows as service rules
                for k in range(svc_header_idx + 1, min(svc_header_idx + 30, len(data))):
                    rule_row = data[k]
                    if not any(rule_row):
                        break
                    svc_name = rule_row[svc_col].strip().upper() if svc_col is not None and svc_col < len(rule_row) else ""
                    if not svc_name:
                        continue

                    etd_day_raw = rule_row[etd_hph_col].strip().upper() if etd_hph_col is not None and etd_hph_col < len(rule_row) else ""
                    cy_raw  = rule_row[cy_col].strip().upper()  if cy_col  is not None and cy_col  < len(rule_row) else ""
                    si_raw  = rule_row[si_col].strip().upper()  if si_col  is not None and si_col  < len(rule_row) else ""

                    # Parse cy_raw like "24:00 SAT" or "SAT 24:00"
                    cy_time, cy_day = _parse_time_day(cy_raw)
                    si_time, si_day = _parse_time_day(si_raw)
                    etd_day = etd_day_raw[:3] if etd_day_raw else ""

                    service_rules[svc_name] = {
                        "etd_day": etd_day,
                        "cy_day": cy_day, "cy_time": cy_time,
                        "si_day": si_day, "si_time": si_time,
                    }
                break

        # ── Parse data rows ───────────────────────────────────────────────
        for i in range(header_row_idx + 1, len(data)):
            row = data[i]
            if not any(row):
                continue

            vessel = row[c_vessel].strip() if c_vessel < len(row) else ""
            voyage = row[c_voyage].strip() if c_voyage < len(row) else ""
            etd_raw = row[c_etd].strip() if c_etd < len(row) else ""
            eta_raw = row[c_eta].strip() if c_eta < len(row) else ""

            if not vessel or not voyage or not etd_raw:
                continue
            if vessel.upper() in ("VESSEL", "VESSEL NAME", ""):
                continue

            # Parse ETD date
            etd_str = safe_date_str(etd_raw)
            if not etd_str:
                continue

            try:
                etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")
            except ValueError:
                continue

            # Filter by date range
            if etd_dt < start_dt or etd_dt > end_dt:
                continue

            # Parse ETA (may have transit time in brackets)
            eta_str = ""
            tt_str = ""
            eta_raw_clean = re.sub(r"\(.*?\)", "", eta_raw).strip()
            eta_str = safe_date_str(eta_raw_clean)

            # Extract T/T from brackets e.g. "(2 days)"
            tt_match = re.search(r"\((\d+)\s*days?\)", eta_raw, re.IGNORECASE)
            if tt_match:
                tt_str = f"{tt_match.group(1)} days"
            elif eta_str and etd_str:
                try:
                    eta_dt = datetime.strptime(eta_str, "%Y/%m/%d")
                    tt_days = (eta_dt - etd_dt).days
                    if tt_days >= 0:
                        tt_str = f"{tt_days} days"
                except Exception:
                    pass

            # Get service name
            svc_name = ""
            if c_service is not None and c_service < len(row):
                svc_name = row[c_service].strip().upper()

            # Calculate CY and SI cut-off
            cy_cutoff = ""
            si_cutoff = ""
            rule = service_rules.get(svc_name, {})
            if rule:
                cy_day  = rule.get("cy_day", "")
                cy_time = rule.get("cy_time", "")
                si_day  = rule.get("si_day", "")
                si_time = rule.get("si_time", "")
                if cy_day:
                    cy_dt = days_back_from_etd(etd_dt, cy_day)
                    cy_cutoff = cy_dt.strftime("%Y/%m/%d") + (f" {cy_time}" if cy_time else "")
                if si_day:
                    si_dt = days_back_from_etd(etd_dt, si_day)
                    si_cutoff = si_dt.strftime("%Y/%m/%d") + (f" {si_time}" if si_time else "")

            # Determine PODs for this row
            pods_in_row = []
            if pod_cols:
                for pod, j in pod_cols.items():
                    if j < len(row) and row[j].strip():
                        pods_in_row.append(pod)
            else:
                # Try to detect POD from row itself
                row_upper = " ".join(row).upper()
                for pod in POD_LIST:
                    if pod in row_upper:
                        pods_in_row.append(pod)

            if not pods_in_row:
                # Default: check if there's explicit POD column
                pods_in_row = [POD_LIST[0]]  # fallback

            for pod in pods_in_row:
                out_rows.append({
                    "POL": "HAIPHONG", "POD": pod,
                    "Vessel": vessel, "Voyage": voyage,
                    "ETD": etd_str, "ETA": eta_str,
                    "T/T Time": tt_str,
                    "CY Cut-off": cy_cutoff,
                    "SI Cut-off": si_cutoff,
                })

    def _parse_time_day(raw: str):
        """Parse strings like '24:00 SAT' or 'SAT 09:00' or '09:00 FRI'"""
        raw = raw.strip().upper()
        time_match = re.search(r"\d{1,2}:\d{2}", raw)
        day_match  = re.search(r"\b(MON|TUE|WED|THU|FRI|SAT|SUN)\b", raw)
        t = time_match.group(0) if time_match else ""
        d = day_match.group(0)[:3]  if day_match  else ""
        return t, d

    if fname_lower.endswith((".xlsx", ".xls")):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            _parse_tsl_excel(wb)
        except Exception as e:
            st.error(f"TSL Excel parse error: {e}")
    elif fname_lower.endswith(".pdf"):
        if not PDF_AVAILABLE:
            st.error("pdfplumber is required to parse PDFs. Run: pip install pdfplumber")
            return empty_df()
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        # Build a minimal worksheet-like structure
                        # Find VESSEL in first few rows
                        header_idx = None
                        for i, row in enumerate(table):
                            if any("VESSEL" in str(c or "").upper() for c in row):
                                header_idx = i
                                break
                        if header_idx is None:
                            continue

                        header = [str(c or "").strip().upper() for c in table[header_idx]]

                        def find_col_pdf(kws):
                            for j, h in enumerate(header):
                                if any(k.upper() in h for k in kws):
                                    return j
                            return None

                        c_v   = find_col_pdf(["VESSEL"])
                        c_voy = find_col_pdf(["VOYAGE"])
                        c_svc = find_col_pdf(["SERVICE"])
                        c_etd_p = find_col_pdf(["ETD"])
                        c_eta_p = find_col_pdf(["ETA"])

                        if c_v is None or c_voy is None or c_etd_p is None:
                            continue

                        for row in table[header_idx + 1:]:
                            def gp(idx):
                                return str(row[idx] or "").strip() if idx is not None and idx < len(row) else ""

                            vessel  = gp(c_v)
                            voyage  = gp(c_voy)
                            etd_raw = gp(c_etd_p)
                            eta_raw = gp(c_eta_p)

                            if not vessel or not etd_raw:
                                continue

                            etd_str = safe_date_str(etd_raw)
                            if not etd_str:
                                continue
                            try:
                                etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")
                            except ValueError:
                                continue

                            if etd_dt < start_date or etd_dt > end_date:
                                continue

                            eta_str = safe_date_str(re.sub(r"\(.*?\)", "", eta_raw).strip())
                            tt_match = re.search(r"\((\d+)\s*days?\)", eta_raw, re.IGNORECASE)
                            tt_str = f"{tt_match.group(1)} days" if tt_match else ""

                            # No service rules for PDF; CY/SI left blank
                            rows.append({
                                "POL": "HAIPHONG", "POD": "HONG KONG",
                                "Vessel": vessel, "Voyage": voyage,
                                "ETD": etd_str, "ETA": eta_str,
                                "T/T Time": tt_str,
                                "CY Cut-off": "",
                                "SI Cut-off": "",
                            })
        except Exception as e:
            st.error(f"TSL PDF parse error: {e}")
    else:
        st.warning(f"TSL unsupported file format: {filename}")

    return pd.DataFrame(rows, columns=OUTPUT_COLS) if rows else empty_df()


# ─────────────────────────────────────────────
# Web Scraping (Selenium)
# ─────────────────────────────────────────────

def get_driver():
    """Create a Selenium headless Chrome driver."""
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service

        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--lang=zh-TW")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36")

        # Try system chromium first (Streamlit Cloud)
        for binary in ["/usr/bin/chromium", "/usr/bin/chromium-browser",
                       "/usr/bin/google-chrome", "/usr/bin/google-chrome-stable"]:
            if os.path.exists(binary):
                options.binary_location = binary
                break

        for driver_path in ["/usr/bin/chromedriver", "/usr/lib/chromium/chromedriver",
                            "/usr/lib/chromium-browser/chromedriver"]:
            if os.path.exists(driver_path):
                service = Service(driver_path)
                return webdriver.Chrome(service=service, options=options)

        # Fallback: webdriver_manager
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            return webdriver.Chrome(service=service, options=options)
        except Exception:
            pass

        # Last resort: no service arg
        return webdriver.Chrome(options=options)

    except Exception as e:
        st.error(f"Failed to start browser driver: {e}\nPlease ensure chromium and chromedriver are installed.")
        return None


def _js_click(driver, element):
    """Click via JavaScript to bypass overlay/interception issues."""
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", element)


def _dismiss_overlays(driver):
    """Try to dismiss cookie banners and modal overlays."""
    try:
        from selenium.webdriver.common.by import By
        dismiss_xpaths = [
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'agree')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'close')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'ok')]",
            "//button[@aria-label='Close' or @aria-label='close']",
            "//div[contains(@class,'modal')]//button",
            "//div[contains(@class,'cookie')]//button",
            "//div[contains(@class,'consent')]//button",
            "//div[contains(@class,'banner')]//button[last()]",
            "//div[contains(@class,'overlay')]//button",
        ]
        for xpath in dismiss_xpaths:
            btns = driver.find_elements(By.XPATH, xpath)
            for btn in btns:
                try:
                    if btn.is_displayed():
                        _js_click(driver, btn)
                        time.sleep(0.4)
                except Exception:
                    pass
    except Exception:
        pass


def _select_flexible(select_el, target: str) -> bool:
    """Try multiple strategies to select an option matching target text."""
    from selenium.webdriver.support.ui import Select
    sel = Select(select_el)
    t = target.strip().upper()
    for opt in sel.options:
        txt = opt.text.strip().upper()
        val = (opt.get_attribute("value") or "").upper()
        if txt == t or t in txt or (val and (t in val or val in t)):
            try:
                sel.select_by_visible_text(opt.text)
                return True
            except Exception:
                try:
                    sel.select_by_value(opt.get_attribute("value"))
                    return True
                except Exception:
                    pass
    return False


# ─────────────────────────────────────────────
# IAL Scraper
# ─────────────────────────────────────────────

def scrape_ial(year: int, month: int, pods: list) -> pd.DataFrame:
    """
    Scrape IAL schedule from https://www.interasia.cc/Service/Form?servicetype=3
    Form: 4 <select> in order: [0] origin country, [1] origin port,
          [2] dest country, [3] dest port.
    Result table columns (Chinese): 出發地 / 目的地 / 出發船名 / 出發航次 /
                                    出發日期 / 抵達日期 / 預估運輸時間
    """
    try:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select, WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
    except ImportError:
        st.error("Please install selenium: pip install selenium")
        return empty_df()

    pod_config = {
        "HONG KONG": ("HONG KONG", "HONG KONG"),
        "SHEKOU":    ("CHINA",     "SHEKOU"),
        "KAOHSIUNG": ("TAIWAN",    "KAOHSIUNG"),
        "TAICHUNG":  ("TAIWAN",    "TAICHUNG"),
    }

    rows = []
    driver = get_driver()
    if driver is None:
        return empty_df()

    url = "https://www.interasia.cc/Service/Form?servicetype=3"

    try:
        for pod in pods:
            if pod not in pod_config:
                continue
            dest_country, dest_port = pod_config[pod]
            try:
                driver.get(url)
                wait = WebDriverWait(driver, 25)
                time.sleep(2)
                _dismiss_overlays(driver)

                wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "select")))
                time.sleep(1)

                selects = driver.find_elements(By.TAG_NAME, "select")
                if len(selects) < 2:
                    st.warning(f"IAL: Found only {len(selects)} select elements on page")
                    continue

                # [0] Origin country = VIETNAM
                for v in ["VIETNAM", "VN", "Vietnam"]:
                    if _select_flexible(selects[0], v):
                        break
                time.sleep(1.5)

                selects = driver.find_elements(By.TAG_NAME, "select")

                # [1] Origin port = HAIPHONG
                if len(selects) >= 2:
                    for v in ["HAIPHONG", "HAI PHONG", "HPH", "Haiphong"]:
                        if _select_flexible(selects[1], v):
                            break
                time.sleep(1.5)

                selects = driver.find_elements(By.TAG_NAME, "select")

                # [2] Dest country
                if len(selects) >= 3:
                    for v in [dest_country, dest_country.split()[0]]:
                        if _select_flexible(selects[2], v):
                            break
                time.sleep(1.5)

                selects = driver.find_elements(By.TAG_NAME, "select")

                # [3] Dest port
                if len(selects) >= 4:
                    for v in [dest_port, dest_port[:4]]:
                        if _select_flexible(selects[3], v):
                            break
                time.sleep(1)

                # Click 查詢 / Search button
                search_btn = None
                for xpath in [
                    "//button[contains(text(),'查詢')]",
                    "//button[contains(text(),'Search')]",
                    "//input[@type='submit']",
                    "//button[@type='submit']",
                    "//a[contains(text(),'查詢')]",
                    "//input[@type='button']",
                ]:
                    els = driver.find_elements(By.XPATH, xpath)
                    if els:
                        search_btn = els[0]
                        break
                if not search_btn:
                    btns = driver.find_elements(By.TAG_NAME, "button")
                    if btns:
                        search_btn = btns[-1]

                if search_btn:
                    _js_click(driver, search_btn)
                    time.sleep(4)
                else:
                    st.warning("IAL: Cannot find search button")
                    continue

                # Parse result table
                tables = driver.find_elements(By.TAG_NAME, "table")
                for tbl in tables:
                    all_trs = tbl.find_elements(By.XPATH, ".//tr")
                    if len(all_trs) < 2:
                        continue
                    header_cells = all_trs[0].find_elements(By.XPATH, ".//th|.//td")
                    headers = [c.text.strip() for c in header_cells]

                    def fci(kws):
                        for j, h in enumerate(headers):
                            if any(k.lower() in h.lower() for k in kws):
                                return j
                        return None

                    c_vessel = fci(["出發船名", "vessel", "ship"])
                    c_voyage = fci(["出發航次", "voyage"])
                    c_etd    = fci(["出發日期", "etd", "departure"])
                    c_eta    = fci(["抵達日期", "eta", "arrival"])
                    c_tt     = fci(["預估運輸", "transit", "t/t"])

                    if c_vessel is None and c_etd is None:
                        continue

                    for tr in all_trs[1:]:
                        tds = tr.find_elements(By.XPATH, ".//td")
                        if not tds:
                            continue

                        def gc(i):
                            return tds[i].text.strip() if i is not None and i < len(tds) else ""

                        vessel  = gc(c_vessel)
                        voyage  = gc(c_voyage)
                        etd_raw = gc(c_etd)
                        eta_raw = gc(c_eta)
                        tt_raw  = gc(c_tt)

                        if not vessel or not etd_raw:
                            continue

                        etd_str = safe_date_str(etd_raw)
                        eta_str = safe_date_str(eta_raw)

                        try:
                            etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")
                            if etd_dt.year != year or etd_dt.month != month:
                                continue
                        except Exception:
                            pass

                        tt_str = tt_raw.strip() if tt_raw else ""
                        if tt_str and re.match(r"^\d+$", tt_str.split()[0]):
                            if "day" not in tt_str.lower():
                                tt_str = f"{tt_str} Days"

                        rows.append({
                            "POL": "HAIPHONG", "POD": pod,
                            "Vessel": vessel, "Voyage": voyage,
                            "ETD": etd_str, "ETA": eta_str,
                            "T/T Time": tt_str,
                            "CY Cut-off": "",
                            "SI Cut-off": "",
                        })

            except Exception as e:
                st.warning(f"IAL scraping error for {pod}: {e}")
                continue

    finally:
        driver.quit()

    return pd.DataFrame(rows, columns=OUTPUT_COLS) if rows else empty_df()


# ─────────────────────────────────────────────
# KMTC Scraper
# ─────────────────────────────────────────────

def scrape_kmtc(year: int, month: int, pods: list) -> pd.DataFrame:
    """
    Scrape KMTC schedule from https://www.ekmtc.com (Vue.js SPA).
    Direct URL params pre-fill the search. After load, click Search,
    then follow detail links and parse: Vessel/Voyage, Departure Date,
    Arrival Date, Total T/T, CY CUT, VGM closing.
    """
    try:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
    except ImportError:
        st.error("Please install selenium: pip install selenium")
        return empty_df()

    pod_params = {
        "HONG KONG": ("HK", "HKG"),
        "SHEKOU":    ("CN", "SKU"),
        "KAOHSIUNG": ("TW", "KHH"),
        "TAICHUNG":  ("TW", "TXG"),
    }

    rows = []
    driver = get_driver()
    if driver is None:
        return empty_df()

    try:
        for pod in pods:
            if pod not in pod_params:
                continue
            dly_ctr, dly_plc = pod_params[pod]
            yyyymm = f"{year:04d}{month:02d}"
            url = (
                f"https://www.ekmtc.com/index.html#/schedule/leg"
                f"?porCtrCd=VN&porPlcCd=HPH"
                f"&dlyCtrCd={dly_ctr}&dlyPlcCd={dly_plc}"
                f"&yyyymm={yyyymm}&loginChk="
            )
            try:
                driver.get(url)
                time.sleep(6)
                _dismiss_overlays(driver)
                time.sleep(1)

                # Click Search button with JS
                for xpath in [
                    "//button[contains(text(),'Search')]",
                    "//button[contains(text(),'SEARCH')]",
                    "//button[@id='searchBtn']",
                    "//button[contains(@class,'btn-search') or contains(@class,'search-btn')]",
                    "//input[@type='button' and contains(@value,'Search')]",
                ]:
                    els = driver.find_elements(By.XPATH, xpath)
                    for el in els:
                        try:
                            if el.is_displayed():
                                _js_click(driver, el)
                                time.sleep(4)
                                break
                        except Exception:
                            pass

                # Parse the listing page first
                _extract_kmtc_table(driver, pod, year, month, rows)
                _parse_kmtc_calendar(driver, pod, year, month, rows)

                # Follow detail links
                detail_links = []
                for link_xpath in [
                    "//a[contains(@href,'detail')]",
                    "//a[contains(@href,'view')]",
                    "//td//a[@href]",
                    "//div[contains(@class,'cal')]//a[@href]",
                ]:
                    detail_links.extend(driver.find_elements(By.XPATH, link_xpath))

                visited = set()
                for link in detail_links[:50]:
                    try:
                        href = link.get_attribute("href")
                        if not href or href in visited or "#" not in href:
                            continue
                        visited.add(href)
                        driver.get(href)
                        time.sleep(2.5)
                        _extract_kmtc_detail(driver, pod, year, month, rows)
                        driver.back()
                        time.sleep(2)
                    except Exception:
                        continue

            except Exception as e:
                st.warning(f"KMTC scraping error for {pod}: {e}")
                continue

    finally:
        driver.quit()

    return pd.DataFrame(rows, columns=OUTPUT_COLS) if rows else empty_df()


def _extract_kmtc_table(driver, pod: str, year: int, month: int, rows: list):
    """Parse KMTC schedule table on current page."""
    try:
        from selenium.webdriver.common.by import By
        tables = driver.find_elements(By.TAG_NAME, "table")
        for table in tables:
            all_trs = table.find_elements(By.XPATH, ".//tr")
            if len(all_trs) < 2:
                continue
            header_cells = all_trs[0].find_elements(By.XPATH, ".//th|.//td")
            headers = [c.text.strip().lower() for c in header_cells]
            if not any("vessel" in h or "departure" in h or "arrival" in h or "etd" in h
                       for h in headers):
                continue

            def fci(kws):
                for j, h in enumerate(headers):
                    if any(k in h for k in kws):
                        return j
                return None

            c_vessel = fci(["vessel", "ship"])
            c_voyage = fci(["voyage"])
            c_etd    = fci(["departure", "etd", "date of dep"])
            c_eta    = fci(["arrival", "eta", "date of arr"])
            c_tt     = fci(["t/t", "transit", "total t"])
            c_cy     = fci(["cy cut", "cy"])
            c_si     = fci(["vgm", "si cut", "si clos"])

            if c_vessel is None or c_etd is None:
                continue

            for tr in all_trs[1:]:
                tds = tr.find_elements(By.XPATH, ".//td")
                def gc(i): return tds[i].text.strip() if i is not None and i < len(tds) else ""
                vessel  = gc(c_vessel)
                voyage  = gc(c_voyage)
                etd_str = safe_date_str(gc(c_etd))
                eta_str = safe_date_str(gc(c_eta))
                tt_str  = gc(c_tt)
                cy_str  = gc(c_cy)
                si_str  = gc(c_si)
                if not vessel or not etd_str:
                    continue
                try:
                    etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")
                    if etd_dt.year != year or etd_dt.month != month:
                        continue
                except Exception:
                    pass
                if tt_str and re.match(r"^\d+$", tt_str.strip()):
                    tt_str = f"{tt_str} Days"
                is_dup = any(r["Vessel"] == vessel and r["Voyage"] == voyage
                             and r["POD"] == pod for r in rows)
                if not is_dup:
                    rows.append({
                        "POL": "HAIPHONG", "POD": pod,
                        "Vessel": vessel, "Voyage": voyage,
                        "ETD": etd_str, "ETA": eta_str,
                        "T/T Time": tt_str,
                        "CY Cut-off": cy_str,
                        "SI Cut-off": si_str,
                    })
    except Exception:
        pass


def _extract_kmtc_detail(driver, pod: str, year: int, month: int, rows: list):
    """Extract schedule from KMTC detail page using label-value pairs."""
    try:
        from selenium.webdriver.common.by import By

        def find_val(label_kws: list) -> str:
            for kw in label_kws:
                kw_lower = kw.lower()
                # th/td sibling pattern
                for xpath in [
                    f"//th[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{kw_lower}')]/following-sibling::td[1]",
                    f"//td[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{kw_lower}')]/following-sibling::td[1]",
                    f"//dt[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{kw_lower}')]/following-sibling::dd[1]",
                    f"//span[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{kw_lower}')]/following-sibling::span[1]",
                ]:
                    els = driver.find_elements(By.XPATH, xpath)
                    if els:
                        return els[0].text.strip()
            return ""

        vv      = find_val(["vessel/voyage", "vessel"])
        etd_val = find_val(["date of departure", "etd"])
        eta_val = find_val(["date of arrival", "eta"])
        tt_val  = find_val(["total t/t", "transit time", "t/t"])
        cy_val  = find_val(["cy cut", "cy closing"])
        si_val  = find_val(["vgm closing", "si cut", "doc cut"])

        vessel, voyage = "", ""
        if vv:
            parts = vv.strip().split()
            if len(parts) >= 2:
                voyage = parts[-1]
                vessel = " ".join(parts[:-1])
            else:
                vessel = vv

        # Text fallback
        if not vessel or not etd_val:
            body = driver.find_element(By.TAG_NAME, "body").text
            if not vv:
                m = re.search(r"Vessel\s*/\s*Voyage[:\s]+([A-Z][A-Z\s]+?)\s+([\w]+)\b",
                               body, re.IGNORECASE)
                if m:
                    vessel = m.group(1).strip()
                    voyage = m.group(2).strip()
            if not etd_val:
                m = re.search(r"Date\s+of\s+Departure[:\s]+([\d]{4}[/\.\-][\d]{1,2}[/\.\-][\d]{1,2})",
                               body, re.IGNORECASE)
                if m: etd_val = m.group(1)
            if not eta_val:
                m = re.search(r"Date\s+of\s+Arrival[:\s]+([\d]{4}[/\.\-][\d]{1,2}[/\.\-][\d]{1,2})",
                               body, re.IGNORECASE)
                if m: eta_val = m.group(1)
            if not tt_val:
                m = re.search(r"Total\s+T/T[:\s]+(\d+\s*Days?)", body, re.IGNORECASE)
                if m: tt_val = m.group(1).strip()
            if not cy_val:
                m = re.search(r"CY\s+CUT[:\s]+([\d]{4}[/\.\-][\d]{1,2}[/\.\-][\d]{1,2}\s*[\d:]*)",
                               body, re.IGNORECASE)
                if m: cy_val = m.group(1).strip()
            if not si_val:
                m = re.search(r"VGM\s+closing[:\s]+([\d]{4}[/\.\-][\d]{1,2}[/\.\-][\d]{1,2}\s*[\d:]*)",
                               body, re.IGNORECASE)
                if m: si_val = m.group(1).strip()

        if not vessel or not etd_val:
            return

        etd_str = safe_date_str(etd_val)
        if not etd_str:
            return
        try:
            etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")
            if etd_dt.year != year or etd_dt.month != month:
                return
        except Exception:
            return

        if tt_val and re.match(r"^\d+$", tt_val.strip()):
            tt_val = f"{tt_val} Days"

        is_dup = any(r["Vessel"] == vessel and r["Voyage"] == voyage
                     and r["POD"] == pod for r in rows)
        if not is_dup:
            rows.append({
                "POL": "HAIPHONG", "POD": pod,
                "Vessel": vessel, "Voyage": voyage,
                "ETD": etd_str, "ETA": safe_date_str(eta_val),
                "T/T Time": tt_val,
                "CY Cut-off": cy_val,
                "SI Cut-off": si_val,
            })
    except Exception:
        pass


def _parse_kmtc_calendar(driver, pod: str, year: int, month: int, rows: list):
    """Parse KMTC calendar-style listing from page body text."""
    try:
        from selenium.webdriver.common.by import By
        body_text = driver.find_element(By.TAG_NAME, "body").text
        lines = body_text.split("\n")
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            date_m = re.match(r"(\d{4})[/\-\.](\d{2})[/\-\.](\d{2})", line)
            if date_m:
                y, mo, d = int(date_m.group(1)), int(date_m.group(2)), int(date_m.group(3))
                if y == year and mo == month:
                    etd_str = f"{y:04d}/{mo:02d}/{d:02d}"
                    vessel, voyage = "", ""
                    for j in range(i + 1, min(i + 10, len(lines))):
                        l2 = lines[j].strip()
                        if re.match(r"^[A-Z][A-Z\s]{3,}$", l2):
                            vessel = l2
                        elif re.match(r"^[A-Z0-9]{3,8}[NnSsEeWw]?$", l2):
                            voyage = l2
                        elif re.match(r"\d{4}[/\-\.]\d{2}[/\-\.]\d{2}", l2):
                            break
                    if vessel:
                        is_dup = any(r["Vessel"] == vessel and r["ETD"] == etd_str
                                     and r["POD"] == pod for r in rows)
                        if not is_dup:
                            rows.append({
                                "POL": "HAIPHONG", "POD": pod,
                                "Vessel": vessel, "Voyage": voyage,
                                "ETD": etd_str, "ETA": "",
                                "T/T Time": "", "CY Cut-off": "", "SI Cut-off": "",
                            })
            i += 1
    except Exception:
        pass


# ─────────────────────────────────────────────
# YML Scraper
# ─────────────────────────────────────────────

def scrape_yml(year: int, month: int, pods: list) -> pd.DataFrame:
    """
    Scrape YML Point-to-Point schedule.
    Key fixes:
    1. _dismiss_overlays() before every click to clear cookie banners
    2. Use _js_click() everywhere to bypass ElementClickInterceptedException
    3. Broad input field discovery via visible inputs list
    4. Autocomplete: try multiple suggestion container selectors
    5. Period: set via JS value assignment
    """
    try:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.keys import Keys
    except ImportError:
        st.error("Please install selenium: pip install selenium")
        return empty_df()

    pod_search_map = {
        "HONG KONG": "Hong Kong",
        "SHEKOU":    "Shekou",
        "KAOHSIUNG": "Kaohsiung",
        "TAICHUNG":  "Taichung",
    }

    rows = []
    base_url = "https://www.yangming.com/en/esolution/schedule/point_to_point_search"
    driver = get_driver()
    if driver is None:
        return empty_df()

    # Pre-compute date range
    import calendar as cal_mod
    last_day = cal_mod.monthrange(year, month)[1]
    date_from = f"{year:04d}/{month:02d}/01"
    date_to   = f"{year:04d}/{month:02d}/{last_day:02d}"

    try:
        for pod in pods:
            if pod not in pod_search_map:
                continue
            pod_kw = pod_search_map[pod]

            try:
                driver.get(base_url)
                wait = WebDriverWait(driver, 25)
                time.sleep(4)
                _dismiss_overlays(driver)
                time.sleep(0.5)

                # ── Locate all visible text inputs ─────────────────────────
                def get_visible_inputs():
                    return [
                        inp for inp in driver.find_elements(
                            By.XPATH,
                            "//input[@type='text' or @type='search' or not(@type)]"
                        ) if inp.is_displayed() and inp.is_enabled()
                    ]

                visible = get_visible_inputs()

                # ── From field: try label-based XPaths, then fallback [0] ──
                from_inp = None
                for xpath in [
                    "//label[contains(text(),'From') or contains(text(),'from')]/following::input[@type!='hidden'][1]",
                    "//span[contains(text(),'From')]/following::input[@type!='hidden'][1]",
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'from')]",
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'location')]",
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'departure')]",
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'origin')]",
                ]:
                    els = driver.find_elements(By.XPATH, xpath)
                    vis = [e for e in els if e.is_displayed() and e.is_enabled()]
                    if vis:
                        from_inp = vis[0]
                        break
                if not from_inp and visible:
                    from_inp = visible[0]

                if from_inp:
                    driver.execute_script("arguments[0].value = '';", from_inp)
                    driver.execute_script("arguments[0].focus();", from_inp)
                    from_inp.send_keys("Haiphong")
                    time.sleep(2)
                    # Accept autocomplete
                    _accept_autocomplete(driver, "haiphong")
                else:
                    st.warning(f"YML: Cannot locate From field for {pod}")

                # ── To field ───────────────────────────────────────────────
                to_inp = None
                for xpath in [
                    "//label[contains(text(),' To') or text()='To']/following::input[@type!='hidden'][1]",
                    "//span[contains(text(),' To') or text()='To']/following::input[@type!='hidden'][1]",
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'to')]",
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'destination')]",
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'arrival')]",
                ]:
                    els = driver.find_elements(By.XPATH, xpath)
                    vis = [e for e in els if e.is_displayed() and e.is_enabled()]
                    if vis:
                        to_inp = vis[0]
                        break
                if not to_inp:
                    visible = get_visible_inputs()
                    if len(visible) >= 2:
                        to_inp = visible[1]

                if to_inp:
                    driver.execute_script("arguments[0].value = '';", to_inp)
                    driver.execute_script("arguments[0].focus();", to_inp)
                    to_inp.send_keys(pod_kw)
                    time.sleep(2)
                    _accept_autocomplete(driver, pod_kw.lower())
                else:
                    st.warning(f"YML: Cannot locate To field for {pod}")

                # ── Period inputs ──────────────────────────────────────────
                period_inputs = driver.find_elements(By.XPATH,
                    "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'period')"
                    " or contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'date')"
                    " or contains(@id,'period') or contains(@name,'period')]"
                )
                if len(period_inputs) >= 2:
                    driver.execute_script(f"arguments[0].value='{date_from}';", period_inputs[0])
                    driver.execute_script(f"arguments[1].value='{date_to}';", period_inputs[1])
                elif len(period_inputs) == 1:
                    driver.execute_script(f"arguments[0].value='{date_from}';", period_inputs[0])

                # ── Search button: dismiss overlays FIRST, then JS click ───
                _dismiss_overlays(driver)
                time.sleep(0.5)

                search_btn = None
                for xpath in [
                    "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'search')]",
                    "//input[@type='submit']",
                    "//button[@type='submit']",
                    "//a[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'search')]",
                    "//button[contains(@class,'search') or contains(@id,'search')]",
                ]:
                    els = driver.find_elements(By.XPATH, xpath)
                    vis = [e for e in els if e.is_displayed()]
                    if vis:
                        search_btn = vis[0]
                        break

                if search_btn:
                    _js_click(driver, search_btn)
                    time.sleep(5)
                else:
                    st.warning(f"YML: Cannot find Search button for {pod}")
                    continue

                # ── Parse results ──────────────────────────────────────────
                # Tables
                tables = driver.find_elements(By.TAG_NAME, "table")
                for tbl in tables:
                    all_trs = tbl.find_elements(By.XPATH, ".//tr")
                    if len(all_trs) < 2:
                        continue
                    header_cells = all_trs[0].find_elements(By.XPATH, ".//th|.//td")
                    headers = [c.text.strip().lower() for c in header_cells]
                    if not any("vessel" in h or "etd" in h or "departure" in h for h in headers):
                        continue

                    def fci(kws):
                        for j, h in enumerate(headers):
                            if any(k in h for k in kws):
                                return j
                        return None

                    c_vessel = fci(["vessel"])
                    c_voyage = fci(["voyage"])
                    c_etd_y  = fci(["etd", "departure"])
                    c_eta_y  = fci(["eta", "arrival"])
                    c_tt     = fci(["t/t", "transit"])
                    c_cy     = fci(["cy"])
                    c_si     = fci(["si cut", "si "])

                    for tr in all_trs[1:]:
                        tds = tr.find_elements(By.XPATH, ".//td")
                        def gc(i): return tds[i].text.strip() if i is not None and i < len(tds) else ""
                        vessel  = gc(c_vessel)
                        voyage  = gc(c_voyage)
                        etd_str = safe_date_str(gc(c_etd_y))
                        eta_str = safe_date_str(gc(c_eta_y))
                        tt_str  = gc(c_tt)
                        cy_str  = gc(c_cy)
                        si_str  = gc(c_si)
                        if not vessel or not etd_str:
                            continue
                        try:
                            etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")
                            if etd_dt.year != year or etd_dt.month != month:
                                continue
                        except Exception:
                            pass
                        if tt_str and re.match(r"^\d+$", tt_str.strip()):
                            tt_str = f"{tt_str} Days"
                        rows.append({
                            "POL": "HAIPHONG", "POD": pod,
                            "Vessel": vessel, "Voyage": voyage,
                            "ETD": etd_str, "ETA": eta_str,
                            "T/T Time": tt_str,
                            "CY Cut-off": cy_str,
                            "SI Cut-off": si_str,
                        })

                # Div/card-based results
                for card_xpath in [
                    "//div[contains(@class,'result') and .//text()[string-length(.) > 5]]",
                    "//div[contains(@class,'schedule')]",
                    "//div[contains(@class,'voyage')]",
                    "//li[contains(@class,'schedule')]",
                ]:
                    for card in driver.find_elements(By.XPATH, card_xpath):
                        _parse_yml_card(card.text, pod, year, month, rows)

            except Exception as e:
                st.warning(f"YML scraping error for {pod}: {e}")
                continue

    finally:
        driver.quit()

    return pd.DataFrame(rows, columns=OUTPUT_COLS) if rows else empty_df()


def _accept_autocomplete(driver, keyword: str):
    """Accept autocomplete suggestion containing keyword."""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    # Try multiple dropdown/suggestion containers
    for ac_xpath in [
        f"//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{keyword}') and (self::li or self::div or self::span)]",
        "//ul[contains(@class,'suggest') or contains(@class,'dropdown') or contains(@class,'auto')]//li[1]",
        "//div[contains(@class,'suggest') or contains(@class,'dropdown') or contains(@class,'option')][1]",
        "//*[contains(@class,'suggestion-item') or contains(@class,'autocomplete-item')][1]",
        "//*[@role='option'][1]",
        "//*[@role='listbox']//li[1]",
    ]:
        suggestions = driver.find_elements(By.XPATH, ac_xpath)
        vis = [s for s in suggestions if s.is_displayed()]
        if vis:
            try:
                _js_click(driver, vis[0])
                time.sleep(0.8)
                return
            except Exception:
                pass
    # Fallback: keyboard navigation
    try:
        active = driver.switch_to.active_element
        active.send_keys(Keys.ARROW_DOWN)
        time.sleep(0.3)
        active.send_keys(Keys.RETURN)
        time.sleep(0.5)
    except Exception:
        pass


def _parse_yml_card(text: str, pod: str, year: int, month: int, rows: list):
    """Parse YML schedule data from a text block."""
    if not text or len(text) < 10:
        return
    date_p = r"(\d{4}[/\-\.]\d{2}[/\-\.]\d{2})"
    dates = re.findall(date_p, text)
    if not dates:
        return

    # Vessel: all-caps sequence
    vessel_m = re.search(r"\b([A-Z][A-Z\s]{3,30})\b", text)
    voyage_m = re.search(r"\b([A-Z0-9]{3,8}[NnSsEeWw]?)\b", text)
    vessel   = vessel_m.group(1).strip() if vessel_m else ""
    voyage   = voyage_m.group(1).strip() if voyage_m else ""

    etd_str = safe_date_str(dates[0]) if dates else ""
    eta_str = safe_date_str(dates[1]) if len(dates) >= 2 else ""

    if not vessel or not etd_str:
        return

    try:
        etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")
        if etd_dt.year != year or etd_dt.month != month:
            return
    except Exception:
        return

    cy_m = re.search(r"CY[:\s]*(" + date_p[1:-1] + r"[\s\d:]*)", text, re.IGNORECASE)
    si_m = re.search(r"(?:SI\s*Cut|Contact)[:\s]*([^\n]+)", text, re.IGNORECASE)
    tt_m = re.search(r"T/T[:\s]*(\d+\s*Days?)", text, re.IGNORECASE)

    is_dup = any(r["Vessel"] == vessel and r["Voyage"] == voyage
                 and r["POD"] == pod for r in rows)
    if not is_dup:
        rows.append({
            "POL": "HAIPHONG", "POD": pod,
            "Vessel": vessel, "Voyage": voyage,
            "ETD": etd_str, "ETA": eta_str,
            "T/T Time": tt_m.group(1).strip() if tt_m else "",
            "CY Cut-off": cy_m.group(1).strip() if cy_m else "",
            "SI Cut-off": si_m.group(1).strip() if si_m else "",
        })



# ─────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────

def export_to_excel(all_data: pd.DataFrame) -> bytes:
    """Create formatted Excel with one sheet per POD, sorted by ETD."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    header_fill  = PatternFill("solid", fgColor="1F4788")
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_fill     = PatternFill("solid", fgColor="DCE6F1")
    normal_font  = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    thin = Side(border_style="thin", color="B8CCE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [12, 12, 20, 10, 12, 12, 10, 20, 22]

    for pod in POD_LIST:
        pod_df = all_data[all_data["POD"] == pod].copy()
        if pod_df.empty:
            continue

        # Sort by ETD
        pod_df["_etd_sort"] = pd.to_datetime(pod_df["ETD"], format="%Y/%m/%d", errors="coerce")
        pod_df = pod_df.sort_values("_etd_sort").drop(columns=["_etd_sort"])
        pod_df = pod_df.reset_index(drop=True)

        # Add carrier column at front
        if "Carrier" in pod_df.columns:
            cols = ["Carrier"] + OUTPUT_COLS
        else:
            cols = OUTPUT_COLS

        ws = wb.create_sheet(title=pod.title())

        # Header row
        ws.row_dimensions[1].height = 30
        for j, col_name in enumerate(OUTPUT_COLS, 1):
            cell = ws.cell(row=1, column=j, value=col_name)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align
            cell.border = border

        # Data rows
        for i, (_, row) in enumerate(pod_df.iterrows(), 2):
            fill = alt_fill if i % 2 == 0 else None
            for j, col_name in enumerate(OUTPUT_COLS, 1):
                val = row.get(col_name, "")
                cell = ws.cell(row=i, column=j, value=str(val) if val else "")
                cell.font   = normal_font
                cell.alignment = center_align
                cell.border = border
                if fill:
                    cell.fill = fill

        # Column widths
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

        # Freeze top row
        ws.freeze_panes = "A2"

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.cell(1, 1, "POD").font = Font(bold=True)
    ws_sum.cell(1, 2, "Count").font = Font(bold=True)
    ws_sum.cell(1, 3, "Carriers").font = Font(bold=True)
    for i, pod in enumerate(POD_LIST, 2):
        subset = all_data[all_data["POD"] == pod]
        ws_sum.cell(i, 1, pod)
        ws_sum.cell(i, 2, len(subset))
        carriers = ", ".join(sorted(subset["Carrier"].unique())) if "Carrier" in subset.columns else "—"
        ws_sum.cell(i, 3, carriers)
    ws_sum.column_dimensions["A"].width = 14
    ws_sum.column_dimensions["B"].width = 8
    ws_sum.column_dimensions["C"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Shipping Schedule Organizer",
        page_icon="🚢",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # Custom CSS
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(90deg, #1F4788 0%, #2E6BC9 100%);
        color: white; padding: 16px 24px; border-radius: 8px;
        margin-bottom: 20px;
    }
    .main-header h1 { margin:0; font-size:1.8rem; }
    .main-header p  { margin:4px 0 0; font-size:0.9rem; opacity:0.85; }
    .carrier-badge {
        display:inline-block; background:#1F4788; color:white;
        padding:2px 10px; border-radius:12px; font-size:0.8rem;
        margin:2px;
    }
    .stat-box {
        background:#F0F4FC; border:1px solid #B8CCE4; border-radius:6px;
        padding:12px; text-align:center;
    }
    .stat-num { font-size:1.8rem; font-weight:700; color:#1F4788; }
    .stat-lbl { font-size:0.8rem; color:#666; }
    div[data-testid="stTabs"] button { font-size:0.95rem; padding:8px 16px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="main-header">
      <h1>🚢 Shipping Schedule Organizer</h1>
      <p>POL: HAIPHONG &rarr; HKG / SKU / KHH / TXG</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Session state ────────────────────────────────────────────────────
    if "data_store" not in st.session_state:
        st.session_state.data_store = {c: empty_df() for c in CARRIER_LIST}

    def store_data(carrier: str, df: pd.DataFrame):
        if not df.empty:
            df["Carrier"] = carrier
        st.session_state.data_store[carrier] = df

    def get_all_data() -> pd.DataFrame:
        frames = []
        for carrier in CARRIER_LIST:
            df = st.session_state.data_store.get(carrier, empty_df())
            if not df.empty:
                if "Carrier" not in df.columns:
                    df["Carrier"] = carrier
                frames.append(df)
        if frames:
            return pd.concat(frames, ignore_index=True)
        return empty_df()

    # ── Sidebar ──────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### ⚙️ Settings")

        # Month & Year selector
        today = datetime.today()
        col_y, col_m = st.columns(2)
        with col_y:
            sel_year  = st.selectbox("Year", list(range(today.year - 1, today.year + 3)),
                                     index=1)
        with col_m:
            sel_month = st.selectbox("Month", list(range(1, 13)),
                                     index=today.month - 1,
                                     format_func=lambda m: f"{m:02d}")

        # Date range for file upload filtering
        import calendar
        last_day = calendar.monthrange(sel_year, sel_month)[1]
        range_start = datetime(sel_year, sel_month, 1)
        range_end   = datetime(sel_year, sel_month, last_day, 23, 59, 59)

        st.divider()
        st.markdown("### 🎯 Destination Port (POD)")
        sel_pods = []
        for pod in POD_LIST:
            if st.checkbox(pod, value=True, key=f"pod_{pod}"):
                sel_pods.append(pod)

        st.divider()
        st.markdown("### 📊 Collected Data")
        for carrier in CARRIER_LIST:
            df = st.session_state.data_store.get(carrier, empty_df())
            cnt = len(df)
            color = "#1F4788" if cnt > 0 else "#aaa"
            st.markdown(f"<span class='carrier-badge' style='background:{color}'>{carrier}: {cnt} records</span>",
                        unsafe_allow_html=True)

        st.divider()
        if st.button("🗑️ Clear All Data", type="secondary", use_container_width=True):
            st.session_state.data_store = {c: empty_df() for c in CARRIER_LIST}
            st.rerun()

    # ── Main tabs ────────────────────────────────────────────────────────
    tab_upload, tab_web, tab_preview, tab_export = st.tabs([
        "📂 Upload Files (CNC / TSL)",
        "🌐 Web Scraping (IAL / KMTC / YML)",
        "📋 Data Preview",
        "📥 Export Excel",
    ])

    # ── Tab 1: File Upload ───────────────────────────────────────────────
    with tab_upload:
        st.subheader("📂 Upload Schedule Files")
        st.info(f"Filtering date range: **{range_start.strftime('%Y/%m/%d')}** to **{range_end.strftime('%Y/%m/%d')}**")

        col1, col2 = st.columns(2)

        # CNC Upload
        with col1:
            st.markdown("#### 🚢 CNC")
            st.caption("Supported formats: CSV / PDF / Excel｜Filename must contain **CNC**")
            cnc_files = st.file_uploader("Select CNC files", type=["csv", "pdf", "xlsx", "xls"],
                                          accept_multiple_files=True, key="cnc_upload",
                                          label_visibility="collapsed")
            if cnc_files:
                if st.button("Parse CNC Files", type="primary", key="parse_cnc"):
                    all_cnc = []
                    with st.spinner("Parsing..."):
                        for f in cnc_files:
                            df = parse_cnc(f.read(), f.name)
                            if not df.empty:
                                all_cnc.append(df)
                    if all_cnc:
                        merged = pd.concat(all_cnc, ignore_index=True)
                        # Filter to selected PODs
                        merged = merged[merged["POD"].isin(sel_pods)]
                        store_data("CNC", merged)
                        st.success(f"✅ CNC parsed successfully: {len(merged)} records")
                        st.dataframe(merged, use_container_width=True)
                    else:
                        st.warning("⚠️ No matching CNC data found")

        # TSL Upload
        with col2:
            st.markdown("#### 🚢 TSL")
            st.caption("Supported formats: Excel / PDF｜Filename must contain **TSL**")
            tsl_files = st.file_uploader("Select TSL files", type=["xlsx", "xls", "pdf"],
                                          accept_multiple_files=True, key="tsl_upload",
                                          label_visibility="collapsed")
            if tsl_files:
                if st.button("Parse TSL Files", type="primary", key="parse_tsl"):
                    all_tsl = []
                    with st.spinner("Parsing..."):
                        for f in tsl_files:
                            df = parse_tsl(f.read(), f.name, range_start, range_end)
                            if not df.empty:
                                all_tsl.append(df)
                    if all_tsl:
                        merged = pd.concat(all_tsl, ignore_index=True)
                        merged = merged[merged["POD"].isin(sel_pods)]
                        store_data("TSL", merged)
                        st.success(f"✅ TSL parsed successfully: {len(merged)} records")
                        st.dataframe(merged, use_container_width=True)
                    else:
                        st.warning("⚠️ No matching TSL data found")

    # ── Tab 2: Web Scraping ──────────────────────────────────────────────
    with tab_web:
        st.subheader("🌐 Web Scraping")
        st.info(f"Target month: **{sel_year} / {sel_month:02d}** ｜ POD: {', '.join(sel_pods) if sel_pods else '(none selected)'}")

        if not sel_pods:
            st.warning("Please select at least one destination port in the sidebar.")
        else:
            col_ial, col_kmtc, col_yml = st.columns(3)

            # IAL
            with col_ial:
                st.markdown("#### 🌐 IAL")
                st.caption("https://www.interasia.cc")
                ial_cnt = len(st.session_state.data_store.get("IAL", empty_df()))
                if ial_cnt > 0:
                    st.success(f"{ial_cnt} records loaded")
                if st.button("Scrape IAL", type="primary", key="scrape_ial", use_container_width=True):
                    with st.spinner(f"Scraping IAL {sel_year}/{sel_month:02d}..."):
                        df = scrape_ial(sel_year, sel_month, sel_pods)
                    if not df.empty:
                        store_data("IAL", df)
                        st.success(f"✅ IAL done: {len(df)} records")
                        st.dataframe(df, use_container_width=True)
                    else:
                        st.warning("⚠️ No IAL data retrieved (please check browser driver installation)")

            # KMTC
            with col_kmtc:
                st.markdown("#### 🌐 KMTC")
                st.caption("https://www.ekmtc.com")
                kmtc_cnt = len(st.session_state.data_store.get("KMTC", empty_df()))
                if kmtc_cnt > 0:
                    st.success(f"{kmtc_cnt} records loaded")
                if st.button("Scrape KMTC", type="primary", key="scrape_kmtc", use_container_width=True):
                    with st.spinner(f"Scraping KMTC {sel_year}/{sel_month:02d}..."):
                        df = scrape_kmtc(sel_year, sel_month, sel_pods)
                    if not df.empty:
                        store_data("KMTC", df)
                        st.success(f"✅ KMTC done: {len(df)} records")
                        st.dataframe(df, use_container_width=True)
                    else:
                        st.warning("⚠️ No KMTC data retrieved (please check browser driver installation)")

            # YML
            with col_yml:
                st.markdown("#### 🌐 YML")
                st.caption("https://www.yangming.com")
                yml_cnt = len(st.session_state.data_store.get("YML", empty_df()))
                if yml_cnt > 0:
                    st.success(f"{yml_cnt} records loaded")
                if st.button("Scrape YML", type="primary", key="scrape_yml", use_container_width=True):
                    with st.spinner(f"Scraping YML {sel_year}/{sel_month:02d}..."):
                        df = scrape_yml(sel_year, sel_month, sel_pods)
                    if not df.empty:
                        store_data("YML", df)
                        st.success(f"✅ YML done: {len(df)} records")
                        st.dataframe(df, use_container_width=True)
                    else:
                        st.warning("⚠️ No YML data retrieved (please check browser driver installation)")

            st.divider()
            st.markdown("#### Scrape All at Once")
            if st.button("🚀 Scrape IAL + KMTC + YML", type="primary", use_container_width=True):
                carriers_web = [("IAL", scrape_ial), ("KMTC", scrape_kmtc), ("YML", scrape_yml)]
                for carrier_name, scrape_fn in carriers_web:
                    with st.spinner(f"Scraping {carrier_name}..."):
                        df = scrape_fn(sel_year, sel_month, sel_pods)
                    if not df.empty:
                        store_data(carrier_name, df)
                        st.success(f"✅ {carrier_name}: {len(df)} records")
                    else:
                        st.warning(f"⚠️ {carrier_name}: No data")

    # ── Tab 3: Data Preview ──────────────────────────────────────────────
    with tab_preview:
        st.subheader("📋 Data Preview")
        all_data = get_all_data()

        if all_data.empty:
            st.info("No data yet. Please upload files or scrape web schedules first.")
        else:
            # Summary stats
            st.markdown("##### 📊 Summary")
            stat_cols = st.columns(len(POD_LIST) + 1)
            with stat_cols[0]:
                st.markdown(f"""
                <div class="stat-box">
                  <div class="stat-num">{len(all_data)}</div>
                  <div class="stat-lbl">Total Records</div>
                </div>""", unsafe_allow_html=True)
            for i, pod in enumerate(POD_LIST, 1):
                cnt = len(all_data[all_data["POD"] == pod])
                with stat_cols[i]:
                    st.markdown(f"""
                    <div class="stat-box">
                      <div class="stat-num">{cnt}</div>
                      <div class="stat-lbl">{pod}</div>
                    </div>""", unsafe_allow_html=True)

            st.divider()

            # Preview by POD
            view_mode = st.radio("View mode", ["By POD", "All combined"], horizontal=True)

            if view_mode == "By POD":
                pod_tabs = st.tabs([p.title() for p in POD_LIST])
                for i, pod in enumerate(POD_LIST):
                    with pod_tabs[i]:
                        pod_df = all_data[all_data["POD"] == pod].copy()
                        if pod_df.empty:
                            st.info(f"No data for {pod}")
                        else:
                            pod_df["_sort"] = pd.to_datetime(pod_df["ETD"], format="%Y/%m/%d", errors="coerce")
                            pod_df = pod_df.sort_values("_sort").drop(columns=["_sort"])
                            display_cols = ["Carrier"] + OUTPUT_COLS if "Carrier" in pod_df.columns else OUTPUT_COLS
                            st.dataframe(pod_df[display_cols], use_container_width=True, height=400)
                            st.caption(f"{len(pod_df)} records")
            else:
                all_data["_sort"] = pd.to_datetime(all_data["ETD"], format="%Y/%m/%d", errors="coerce")
                all_data = all_data.sort_values(["POD", "_sort"]).drop(columns=["_sort"])
                display_cols = ["Carrier"] + OUTPUT_COLS if "Carrier" in all_data.columns else OUTPUT_COLS
                st.dataframe(all_data[display_cols], use_container_width=True, height=500)

    # ── Tab 4: Export ────────────────────────────────────────────────────
    with tab_export:
        st.subheader("📥 Export Excel")
        all_data = get_all_data()

        if all_data.empty:
            st.warning("⚠️ No data to export. Please collect schedule data first.")
        else:
            st.markdown(f"""
            **Export Summary:**
            - Total records: **{len(all_data)}**
            - Carriers: {', '.join(sorted(all_data['Carrier'].unique()) if 'Carrier' in all_data.columns else [])}
            - POD Sheets: {', '.join([p for p in POD_LIST if p in all_data['POD'].values])}
            """)

            filename = f"ShippingSchedule_{sel_year}{sel_month:02d}.xlsx"
            if st.button("🔄 Generate Excel Report", type="primary"):
                with st.spinner("Generating..."):
                    excel_bytes = export_to_excel(all_data)
                st.success("✅ Excel report ready!")
                st.download_button(
                    label=f"⬇️ Download {filename}",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

            st.divider()
            st.markdown("#### Sheet Preview")
            for pod in POD_LIST:
                pod_df = all_data[all_data["POD"] == pod]
                if not pod_df.empty:
                    with st.expander(f"📄 {pod.title()} ({len(pod_df)} records)"):
                        pod_df2 = pod_df.copy()
                        pod_df2["_s"] = pd.to_datetime(pod_df2["ETD"], format="%Y/%m/%d", errors="coerce")
                        pod_df2 = pod_df2.sort_values("_s").drop(columns=["_s"])
                        display_cols = ["Carrier"] + OUTPUT_COLS if "Carrier" in pod_df2.columns else OUTPUT_COLS
                        st.dataframe(pod_df2[display_cols], use_container_width=True)


if __name__ == "__main__":
    main()
